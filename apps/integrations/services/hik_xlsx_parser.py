"""Parse Hik Connect XLSX/CSV exports into event rows for HikEvent import."""

from __future__ import annotations

import csv
import hashlib
import io
import logging
import re
import zipfile
from datetime import date, datetime
from pathlib import Path

from django.utils import timezone
from django.utils.dateparse import parse_datetime

logger = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None  # type: ignore

CODE_HEADERS = (
    "personcode",
    "personno",
    "employeeno",
    "employeenumber",
    "cardno",
    "cardnumber",
    "personid",
    "id",
    "номеркарты",
    "номеркартыaccess",
    "табельныйномер",
    "номерсотрудника",
    "код",
)
TIME_HEADERS = (
    "time",
    "datetime",
    "date",
    "eventtime",
    "happentime",
    "checktime",
    "attendancetime",
    "время",
    "дата",
    "датаивремя",
    "времяпрохода",
    "времясобытия",
)
DOOR_HEADERS = (
    "door",
    "doorname",
    "device",
    "devicename",
    "accesspoint",
    "checkpoint",
    "location",
    "точкадоступа",
    "устройство",
    "турникет",
    "дверь",
)
TYPE_HEADERS = (
    "eventtype",
    "type",
    "eventname",
    "status",
    "attendancestatus",
    "тип",
    "типсобытия",
    "статус",
)
NAME_HEADERS = ("personname", "name", "employeename", "фио", "имя", "сотрудник")

# Короткие кандидаты ("id", "код", "type") при подстрочном поиске цепляют чужие
# колонки вроде «Device ID» или «Код подразделения», поэтому подстрока
# разрешена только достаточно длинным вариантам.
_MIN_SUBSTRING_CANDIDATE = 5


class HikExportFormatError(ValueError):
    """Выгрузка не распознана: другая структура файла или не тот отчёт.

    Отдельный тип нужен, чтобы задача отличала «данных за день нет» от
    «мы не поняли файл» — раньше оба случая давали пустой список и успех.
    """


def _norm_header(value: str) -> str:
    return re.sub(r"[\s_\-./\\()]+", "", (value or "").strip().lower())


def _pick_column(headers: list[str], candidates: tuple[str, ...]) -> int | None:
    """Найти индекс колонки по списку кандидатов.

    Порядок проверок — от точного к приблизительному, чтобы «Номер карты»
    выигрывал у «Код подразделения», а «Device ID» не считался колонкой кода.
    """
    normalized = [_norm_header(h) for h in headers]

    for idx, header in enumerate(normalized):
        if header and header in candidates:
            return idx

    for idx, header in enumerate(normalized):
        if not header:
            continue
        for cand in candidates:
            if len(cand) >= _MIN_SUBSTRING_CANDIDATE and cand in header:
                return idx

    for idx, header in enumerate(normalized):
        if len(header) < _MIN_SUBSTRING_CANDIDATE:
            continue
        for cand in candidates:
            if header in cand:
                return idx
    return None


def _cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if timezone.is_naive(value):
            return timezone.make_aware(value).isoformat()
        return value.isoformat()
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time()).date().isoformat()
    return str(value).strip()


def _parse_time_value(raw: str, fallback_date: date | None = None) -> str | None:
    """Привести значение ячейки времени к ISO-строке с таймзоной.

    Возвращает None, если значение распознать не удалось. Раньше здесь
    возвращался исходный текст, который ниже по течению превращался
    в `timezone.now()` — время прохода молча подменялось временем импорта,
    и все расчёты опозданий по такому дню были мусором.

    Дата без времени тоже считается нераспознанной: полночь означала бы,
    что студент всегда пришёл вовремя.
    """
    text = (raw or "").strip()
    if not text:
        if fallback_date:
            return timezone.make_aware(datetime.combine(fallback_date, datetime.min.time())).isoformat()
        return None

    # parse_datetime опирается на datetime.fromisoformat, а тот принимает
    # "2026-05-30" и достраивает полночь. Для проходов через турникет это
    # означало бы «пришёл в 00:00», то есть всегда вовремя.
    if ":" not in text:
        return None

    dt = parse_datetime(text.replace("Z", "+00:00"))
    if dt:
        if timezone.is_naive(dt):
            dt = timezone.make_aware(dt, timezone.get_current_timezone())
        return dt.isoformat()

    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y %H:%M",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
    ):
        try:
            dt = datetime.strptime(text, fmt)
            return timezone.make_aware(dt, timezone.get_current_timezone()).isoformat()
        except ValueError:
            continue

    if fallback_date:
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                t = datetime.strptime(text, fmt).time()
                dt = datetime.combine(fallback_date, t)
                return timezone.make_aware(dt, timezone.get_current_timezone()).isoformat()
            except ValueError:
                continue
    return None


def build_export_event_id(
    code: str,
    event_time: str,
    door: str,
    occurrences: dict[str, int] | None = None,
) -> str:
    """Построить стабильный идентификатор события из полей выгрузки.

    Идентификатор зависит только от содержимого события. Раньше в хеш входил
    порядковый номер строки в файле, а выгрузка «за сегодня» скачивается
    каждый час и растёт: если портал отдаёт новые записи сверху, номера строк
    всех прошлых событий сдвигались, они получали новые id и импортировались
    повторно — вместе с повторными штрафами рейтинга.

    `occurrences` считает полные дубликаты внутри одной выгрузки, чтобы два
    реальных прохода с одинаковыми полями не схлопнулись в одно событие.
    Индекс повтора воспроизводим при повторном разборе того же файла.
    """
    base = f"{code}|{event_time}|{door}"
    index = 0
    if occurrences is not None:
        index = occurrences.get(base, 0)
        occurrences[base] = index + 1
    payload = base if index == 0 else f"{base}|dup{index}"
    digest = hashlib.sha1(payload.encode()).hexdigest()[:16]
    return f"xlsx-{digest}"


def _rows_from_xlsx(path: Path) -> list[list[str]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is not installed")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([_cell_str(c) for c in row])
    wb.close()
    return rows


def _rows_from_csv(path: Path) -> list[list[str]]:
    text = path.read_text(encoding="utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    return [[cell.strip() for cell in row] for row in reader]


def _rows_from_zip(path: Path) -> list[list[str]]:
    with zipfile.ZipFile(path) as zf:
        for name in zf.namelist():
            lower = name.lower()
            if lower.endswith(".xlsx"):
                data = zf.read(name)
                tmp = path.with_suffix(".inner.xlsx")
                tmp.write_bytes(data)
                try:
                    return _rows_from_xlsx(tmp)
                finally:
                    tmp.unlink(missing_ok=True)
            if lower.endswith(".csv"):
                text = zf.read(name).decode("utf-8-sig", errors="replace")
                reader = csv.reader(io.StringIO(text))
                return [[cell.strip() for cell in row] for row in reader]
    raise ValueError(f"No xlsx/csv found inside zip: {path}")


def load_tabular_rows(path: str | Path) -> list[list[str]]:
    p = Path(path)
    suffix = p.suffix.lower()
    if suffix == ".zip":
        return _rows_from_zip(p)
    if suffix in {".xlsx", ".xlsm"}:
        return _rows_from_xlsx(p)
    if suffix == ".csv":
        return _rows_from_csv(p)
    raise ValueError(f"Unsupported export format: {suffix}")


def _find_header_row(rows: list[list[str]]) -> tuple[int, int]:
    """Вернуть (индекс строки заголовков, оценка совпадения).

    Оценка нужна вызывающему коду: ноль означает, что строку заголовков найти
    не удалось. Раньше в этом случае молча бралась строка 0 — то есть шапка
    отчёта («Записи прохода», «Период: …») становилась заголовками таблицы.
    """
    best_idx = 0
    best_score = 0
    for idx, row in enumerate(rows[:20]):
        score = 0
        joined = " ".join(_norm_header(c) for c in row if c)
        for group in (CODE_HEADERS, TIME_HEADERS, DOOR_HEADERS, TYPE_HEADERS, NAME_HEADERS):
            if any(token in joined for token in group):
                score += 1
        if score > best_score:
            best_score = score
            best_idx = idx
    return best_idx, best_score


def parse_hik_export_rows(
    path: str | Path,
    *,
    fallback_date: date | None = None,
) -> list[dict]:
    """
    Convert XLSX/CSV/ZIP export to dict rows compatible with save_hik_row_as_event().
    """
    rows = load_tabular_rows(path)
    non_empty = [r for r in rows if any(cell.strip() for cell in r)]
    if not non_empty:
        raise HikExportFormatError(f"Файл выгрузки пуст: {Path(path).name}")

    header_idx, header_score = _find_header_row(non_empty)
    if not header_score:
        raise HikExportFormatError(
            f"В файле {Path(path).name} не найдена строка заголовков — "
            "вероятно, скачан не тот отчёт или изменился формат выгрузки"
        )

    headers = non_empty[header_idx]
    data_rows = non_empty[header_idx + 1 :]

    code_col = _pick_column(headers, CODE_HEADERS)
    time_col = _pick_column(headers, TIME_HEADERS)
    door_col = _pick_column(headers, DOOR_HEADERS)
    type_col = _pick_column(headers, TYPE_HEADERS)
    name_col = _pick_column(headers, NAME_HEADERS)

    if code_col is None and name_col is None:
        raise HikExportFormatError(
            f"В файле {Path(path).name} нет колонки с кодом карты или ФИО. "
            f"Найденные заголовки: {[h for h in headers if h.strip()]}"
        )
    if time_col is None:
        raise HikExportFormatError(
            f"В файле {Path(path).name} нет колонки времени прохода. "
            f"Найденные заголовки: {[h for h in headers if h.strip()]}"
        )

    events: list[dict] = []
    # Счётчик повторов нужен, чтобы два одинаковых прохода (тот же человек,
    # та же секунда, тот же турникет) получили разные, но воспроизводимые id.
    occurrences: dict[str, int] = {}
    skipped_no_code = 0
    skipped_bad_time = 0

    for row in data_rows:
        if not any(cell.strip() for cell in row):
            continue
        code = row[code_col].strip() if code_col is not None and code_col < len(row) else ""
        if not code and name_col is not None and name_col < len(row):
            code = row[name_col].strip()
        if not code:
            skipped_no_code += 1
            continue

        time_raw = row[time_col].strip() if time_col is not None and time_col < len(row) else ""
        event_time = _parse_time_value(time_raw, fallback_date=fallback_date)
        if not event_time:
            skipped_bad_time += 1
            continue

        door = row[door_col].strip() if door_col is not None and door_col < len(row) else ""
        event_type = row[type_col].strip() if type_col is not None and type_col < len(row) else "access"
        person_name = row[name_col].strip() if name_col is not None and name_col < len(row) else ""

        event_id = build_export_event_id(code, event_time, door, occurrences)
        event = {
            "eventId": event_id,
            "personCode": code,
            "eventTime": event_time,
            "eventType": event_type or "access",
            "doorName": door,
        }
        if person_name:
            event["personName"] = person_name
        events.append(event)

    if data_rows and not events:
        raise HikExportFormatError(
            f"В файле {Path(path).name} не удалось разобрать ни одной строки "
            f"(строк данных: {len(data_rows)}, без кода: {skipped_no_code}, "
            f"с нераспознанным временем: {skipped_bad_time})"
        )

    if skipped_no_code or skipped_bad_time:
        logger.warning(
            "hik export %s: пропущено строк — без кода: %s, с нераспознанным временем: %s",
            Path(path).name,
            skipped_no_code,
            skipped_bad_time,
        )

    return events


def snapshot_payload_from_export(path: str | Path, target_date: date) -> dict:
    events = parse_hik_export_rows(path, fallback_date=target_date)
    return {
        "date": target_date.isoformat(),
        "meta": {
            "source": "browser_xlsx",
            "file": Path(path).name,
            "events_count": len(events),
            "imported_at": timezone.now().isoformat(),
        },
        "events": events,
    }
